#!/Library/Frameworks/Python.framework/Versions/3.7/bin/python3
# -*- coding: utf-8 -*-

import subprocess
import datetime
from openpyxl import Workbook
from openpyxl import load_workbook

#시간 설정
now = datetime.datetime.now()
nowdate = now.strftime('%Y-%m-%d')
nowtime = now.strftime('%H:%M:%S')
nowdatetime = now.strftime('%Y-%m-%d %H:%M:%S')

#EC2 infra 정보 / 리턴값
instance_info = "aws ec2 describe-instances --query 'Reservations[].Instances[].[[Tags[?Key==`Name`].Value][0][0], InstanceId, InstanceType, PrivateIpAddress, PublicIpAddress, VpcId, SubnetId, SecurityGroups[].GroupName[],SecurityGroups[].GroupId[], BlockDeviceMappings[].Ebs[].VolumeId[]]' --output text"
instance_return = subprocess.check_output(instance_info, shell=True)
vpc_info = "aws ec2 describe-vpcs --query 'Vpcs[].[CidrBlock, VpcId]' --output text"
vpc_return = subprocess.check_output(vpc_info, shell=True)
subnet_info = "aws ec2 describe-subnets --query 'Subnets[].[AvailabilityZone, CidrBlock, SubnetId, VpcId, SubnetArn]' --output text"
subnet_return = subprocess.check_output(subnet_info, shell=True)
routetable_info = "aws ec2 describe-route-tables --query 'RouteTables[].[[Tags[?Key==`Name`].Value][0][0], RouteTableId, Routes, Associations[].SubnetId[]]' --output text"
routetable_return = subprocess.check_output(routetable_info, shell=True)
internetgw_info = "aws ec2 describe-internet-gateways --query 'InternetGateways[].[[Tags[?Key==`Name`].Value][0][0], Attachments[].VpcId[], InternetGatewayId]' --output text"
internetgw_return = subprocess.check_output(internetgw_info, shell=True)

wb = Workbook() #워크북 생성
ws1 = wb.active #워크 시트 생성
ws1.title = nowdate #시트 이름 수정
ws2 = wb.create_sheet(nowdate) #시트 추가
ws1["A1"] = instance_return
ws1["A2"] = vpc_return
ws1["A3"] = subnet_return
ws1["A4"] = routetable_return
ws1["A5"] = internetgw_return

wb.save("ec2-info-test.xlsx")
exit()

#파일 경로 지정
filename = load_workbook("ec2-info.xlsx")