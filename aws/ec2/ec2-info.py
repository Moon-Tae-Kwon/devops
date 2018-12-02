#!/Library/Frameworks/Python.framework/Versions/3.7/bin/python3
# -*- coding: utf-8 -*-

import os
import datetime
from openpyxl import Workbook
from openpyxl import load_workbook

wb = Workbook()
ws1 = wb.active
ws1.title = "first_1"
ws1["A1"] = "Hello World"
print(ws1["A1"].value)
wb.save("ec2-info-test.xlsx")
exit()

#파일 경로 지정
filename = load_workbook("ec2-info.xlsx")

#시간 설정
now = datetime.datetime.now()
nowdate = now.strftime('%Y-%m-%d')
nowtime = now.strftime('%H:%M:%S')
nowdatetime = now.strftime('%Y-%m-%d %H:%M:%S')

#엑셀 파일 시트 추가
sheet = filename.create_sheet(title = nowdate)


filename.save("ec2-info.xlsx")